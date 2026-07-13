from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import numpy as np
import openpyxl
import pandas as pd


@dataclass(frozen=True)
class CyclicLawParameters:
    """Empirical residual-settlement envelope calibrated against Qiu et al. (2025)."""

    family: str
    coefficients: tuple[float, ...]
    cv_rmse_mm: float
    cv_se_mm: float
    calibration_rmse_mm: float
    interpretation: str = "empirical screening envelope; not a constitutive law"


def load_qiu_fig21(path: str | Path) -> pd.DataFrame:
    """Read all five load-level pairs from the authors' Fig. 21 workbook."""
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    records: list[dict[str, float]] = []
    for start_col in range(1, ws.max_column + 1, 2):
        label = ws.cell(3, start_col + 1).value
        if label is None:
            continue
        load_ratio = float(str(label).split("%")[0]) / 100.0
        for row in range(4, ws.max_row + 1):
            n = ws.cell(row, start_col).value
            s = ws.cell(row, start_col + 1).value
            if n is None or s is None:
                continue
            records.append(
                {
                    "source": "Qiu et al. (2025), Fig. 21 authors' workbook",
                    "load_ratio": load_ratio,
                    "cycle": float(n),
                    "settlement_mm": float(s),
                }
            )
    data = pd.DataFrame(records).sort_values(["load_ratio", "cycle"]).reset_index(drop=True)
    counts = data.groupby("load_ratio")["cycle"].count().to_dict()
    expected = {0.33, 0.50, 0.60, 0.70, 0.80}
    if set(counts) != expected or any(v != 31 for v in counts.values()):
        raise ValueError(f"Expected five Qiu load levels with 31 points each; found {counts}")
    return data


def _logarithmic(x: tuple[np.ndarray, np.ndarray], a: float, b: float) -> np.ndarray:
    r, n = x
    return a * np.power(r, b) * np.log1p(n)


def _power(x: tuple[np.ndarray, np.ndarray], a: float, b: float, c: float) -> np.ndarray:
    r, n = x
    return a * np.power(r, b) * np.power(np.maximum(n, 0.0), c)


def _asymptotic(x: tuple[np.ndarray, np.ndarray], a: float, b: float, k: float) -> np.ndarray:
    r, n = x
    return a * np.power(r, b) * (1.0 - np.exp(-k * np.maximum(n, 0.0)))


FAMILIES: dict[str, tuple[Callable, tuple[float, ...], tuple[tuple[float, ...], tuple[float, ...]]]] = {
    "logarithmic": (_logarithmic, (1.0, 2.0), ((0.0, 0.0), (50.0, 12.0))),
    "power": (_power, (2.0, 3.0, 0.3), ((0.0, 0.0, 0.02), (100.0, 15.0, 2.0))),
    "asymptotic": (_asymptotic, (3.0, 4.0, 0.3), ((0.0, 0.0, 1e-4), (100.0, 15.0, 10.0))),
}


def _positive_log_linear_fit(response: np.ndarray, r: np.ndarray, extra: np.ndarray | None = None) -> np.ndarray:
    columns = [np.ones_like(r), np.log(r)]
    if extra is not None:
        columns.append(np.log(extra))
    beta = np.linalg.lstsq(np.column_stack(columns), np.log(response), rcond=None)[0]
    return beta


def _fit_family(data: pd.DataFrame, family: str) -> tuple[np.ndarray, np.ndarray]:
    """Deterministic constrained fits without a proprietary or heavy solver dependency."""
    positive = data[(data["cycle"] > 0.0) & (data["settlement_mm"] > 0.0)]
    r = positive["load_ratio"].to_numpy(float)
    n = positive["cycle"].to_numpy(float)
    y = positive["settlement_mm"].to_numpy(float)
    if family == "logarithmic":
        beta = _positive_log_linear_fit(y / np.log1p(n), r)
        params = np.array([np.exp(beta[0]), max(0.0, beta[1])])
    elif family == "power":
        beta = _positive_log_linear_fit(y, r, n)
        params = np.array([np.exp(beta[0]), max(0.0, beta[1]), np.clip(beta[2], 0.02, 2.0)])
    elif family == "asymptotic":
        candidates: list[tuple[float, np.ndarray]] = []
        for k in np.geomspace(1e-3, 10.0, 600):
            shape = 1.0 - np.exp(-k * n)
            beta = _positive_log_linear_fit(y / shape, r)
            params_k = np.array([np.exp(beta[0]), max(0.0, beta[1]), k])
            pred = _asymptotic((r, n), *params_k)
            candidates.append((float(np.mean((pred - y) ** 2)), params_k))
        params = min(candidates, key=lambda item: item[0])[1]
    else:
        raise KeyError(family)
    return params, np.empty((len(params), len(params)))


def predict_cyclic_settlement(
    load_ratio: np.ndarray | float,
    completed_cycles: np.ndarray | float,
    parameters: CyclicLawParameters,
) -> np.ndarray:
    fn = FAMILIES[parameters.family][0]
    r = np.asarray(load_ratio, dtype=float)
    n = np.asarray(completed_cycles, dtype=float)
    return fn((r, n), *parameters.coefficients)


def fit_and_select_cyclic_law(data: pd.DataFrame) -> tuple[CyclicLawParameters, pd.DataFrame]:
    """Select by leave-one-load-level-out RMSE with a one-SE simplicity rule."""
    levels = sorted(data["load_ratio"].unique())
    rows: list[dict[str, float | str]] = []
    complexity = {"logarithmic": 2, "power": 3, "asymptotic": 3}
    for family, (fn, _, _) in FAMILIES.items():
        fold_errors: list[float] = []
        for level in levels:
            train = data[data["load_ratio"] != level]
            test = data[data["load_ratio"] == level]
            params, _ = _fit_family(train, family)
            pred = fn(
                (test["load_ratio"].to_numpy(float), test["cycle"].to_numpy(float)),
                *params,
            )
            fold_errors.append(float(np.sqrt(np.mean((pred - test["settlement_mm"].to_numpy(float)) ** 2))))
        rows.append(
            {
                "family": family,
                "parameter_count": complexity[family],
                "cv_rmse_mm": float(np.mean(fold_errors)),
                "cv_se_mm": float(np.std(fold_errors, ddof=1) / np.sqrt(len(fold_errors))),
            }
        )
    comparison = pd.DataFrame(rows).sort_values("cv_rmse_mm").reset_index(drop=True)
    best = comparison.iloc[0]
    eligible = comparison[comparison["cv_rmse_mm"] <= best["cv_rmse_mm"] + best["cv_se_mm"]]
    chosen_row = eligible.sort_values(["parameter_count", "cv_rmse_mm", "family"]).iloc[0]
    family = str(chosen_row["family"])
    params, _ = _fit_family(data, family)
    fn = FAMILIES[family][0]
    fitted = fn((data["load_ratio"].to_numpy(float), data["cycle"].to_numpy(float)), *params)
    rmse = float(np.sqrt(np.mean((fitted - data["settlement_mm"].to_numpy(float)) ** 2)))
    selected = CyclicLawParameters(
        family=family,
        coefficients=tuple(float(v) for v in params),
        cv_rmse_mm=float(chosen_row["cv_rmse_mm"]),
        cv_se_mm=float(chosen_row["cv_se_mm"]),
        calibration_rmse_mm=rmse,
    )
    comparison["selected"] = comparison["family"].eq(family)
    return selected, comparison


def residual_bootstrap_predictions(
    data: pd.DataFrame,
    selected: CyclicLawParameters,
    grid: pd.DataFrame,
    n_bootstrap: int = 300,
    seed: int = 20260713,
) -> pd.DataFrame:
    """Residual bootstrap intervals, preserving the experimental design points."""
    rng = np.random.default_rng(seed)
    fn = FAMILIES[selected.family][0]
    x = (data["load_ratio"].to_numpy(float), data["cycle"].to_numpy(float))
    y = data["settlement_mm"].to_numpy(float)
    fitted = fn(x, *selected.coefficients)
    residuals = y - fitted
    gx = (grid["load_ratio"].to_numpy(float), grid["cycle"].to_numpy(float))
    draws: list[np.ndarray] = []
    for _ in range(n_bootstrap):
        y_star = fitted + rng.choice(residuals, size=len(residuals), replace=True)
        boot = data.copy()
        boot["settlement_mm"] = np.maximum(0.0, y_star)
        try:
            params, _ = _fit_family(boot, selected.family)
        except (RuntimeError, ValueError):
            continue
        draws.append(fn(gx, *params))
    if len(draws) < max(50, n_bootstrap // 2):
        raise RuntimeError("Too few successful bootstrap fits")
    array = np.vstack(draws)
    result = grid.copy()
    result["predicted_mm"] = fn(gx, *selected.coefficients)
    result["lower95_mm"] = np.percentile(array, 2.5, axis=0)
    result["upper95_mm"] = np.percentile(array, 97.5, axis=0)
    result["bootstrap_successes"] = len(draws)
    return result
